# %%
from cgi import test
import re
import pandas as pd
import numpy as np
import pickle
from openpyxl import load_workbook


from sklearn.feature_extraction.text import CountVectorizer
from sklearn import svm
from sklearn.linear_model import SGDClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.naive_bayes import BernoulliNB
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score, precision_recall_curve, confusion_matrix
from sklearn.pipeline import Pipeline

from NPF.teachers_and_covid import start
from NPF.library import classify

FILE_NAME = start.CLEAN_DIR + "relevance model statistics.txt"
f = open(FILE_NAME, "w+")

performance_statistics = {}
# %%
tweets = pd.read_csv(start.CLEAN_DIR + "tweets_full.csv")
tweets = tweets[
    [
        "unique_id",
        "text",
        "created",
        "likes",
        "retweets",
        "quotes",
        "replies",
        "author_id",
        "geo",
        "random_set",
    ]
]
annotations = pd.read_csv(start.CLEAN_DIR + "annotations.csv")
annotations = annotations[
    [
        "unique_id",
        "tweet_id",
        "random_set",
        "relevant",
        "category",
        "covid",
        "character",
    ]
]

df = annotations.merge(
    tweets[["unique_id", "text"]],
    how="left",
    left_on="unique_id",
    right_on="unique_id",
)

# %%
df = df.sample(len(annotations), random_state=68)
testing = df[df.random_set == 3]
training = df[df.random_set != 3]

# %% SVM
pipeline = Pipeline(
    [
        (
            "vect",
            CountVectorizer(
                strip_accents="unicode",
                stop_words=classify.stop_words,
                ngram_range=(1, 3),
                min_df=10,
            ),
        ),
        ("clf", svm.LinearSVC()),
    ],
)
clf = pipeline.fit(training.text, training.relevant)
pickle.dump(clf, open(start.TEMP_DIR + "model_svm.sav", "wb"))


testing["classification_svm"] = clf.predict(testing.text)
testing["score_svm"] = clf.decision_function(testing.text)

training["classification_csv"] = clf.predict(training.text)
training["score_svm"] = clf.decision_function(training.text)

performance_statistics["SVM"] = classify.return_statistics(
    ground_truth=testing.relevant,
    scores=testing.score_svm,
    classification=testing.classification_svm,
)

# %%


# %% Stochastic Gradient Descent
pipeline = Pipeline(
    [
        (
            "vect",
            CountVectorizer(
                strip_accents="unicode",
                stop_words=classify.stop_words,
                ngram_range=(1, 3),
                min_df=10,
            ),
        ),
        ("clf", SGDClassifier(random_state=87)),
    ],
)
clf = pipeline.fit(training.text, training.relevant)

testing["classification_sgd"] = clf.predict(testing.text)
testing["score_sgd"] = clf.decision_function(testing.text)

training["classification_sgd"] = clf.predict(training.text)
training["score_sgd"] = clf.decision_function(training.text)


pickle.dump(clf, open(start.TEMP_DIR + "model_sgd.sav", "wb"))

performance_statistics["SGD"] = classify.return_statistics(
    ground_truth=testing.relevant,
    scores=testing.score_sgd,
    classification=testing.classification_sgd,
)

# %% Random Forest
pipeline = Pipeline(
    [
        (
            "vect",
            CountVectorizer(
                strip_accents="unicode",
                stop_words=classify.stop_words,
                ngram_range=(1, 3),
                min_df=10,
            ),
        ),
        ("clf", RandomForestClassifier(random_state=2)),
    ],
)

clf = pipeline.fit(training.text, training.relevant)

testing["classification_rf"] = clf.predict(testing.text)
testing["score_rf"] = [proba[1] for proba in clf.predict_proba(testing.text)]

training["classification_rf"] = clf.predict(training.text)
training["score_rf"] = [proba[1] for proba in clf.predict_proba(training.text)]

pickle.dump(clf, open(start.TEMP_DIR + "model_rf.sav", "wb"))

performance_statistics["Random Forest"] = classify.return_statistics(
    ground_truth=testing.relevant,
    scores=testing.score_rf,
    classification=testing.classification_rf,
)

# %% Bernoulli Naive Bayes
pipeline = Pipeline(
    [
        (
            "vect",
            CountVectorizer(
                strip_accents="unicode",
                stop_words=classify.stop_words,
                ngram_range=(1, 3),
                min_df=10,
            ),
        ),
        ("clf", BernoulliNB()),
    ],
)

clf = pipeline.fit(training.text, training.relevant)

testing["classification_nb"] = clf.predict(testing.text)
testing["score_nb"] = [proba[1] for proba in clf.predict_proba(testing.text)]


training["classification_nb"] = clf.predict(training.text)
training["score_nb"] = [proba[1] for proba in clf.predict_proba(training.text)]


pickle.dump(clf, open(start.TEMP_DIR + "model_nb.sav", "wb"))

performance_statistics["Naive Bayes"] = classify.return_statistics(
    ground_truth=testing.relevant,
    scores=testing.score_nb,
    classification=testing.classification_nb,
)

# %% Naive Bayes Threshold

probas = [proba[1] for proba in clf.predict_proba(training.text)]
precisions, recalls, thresholds = precision_recall_curve(training.relevant, probas)
threshold_recall = thresholds[np.argmax(recalls >= 0.70)]

testing["classification_nb_recall"] = [
    proba[1] for proba in clf.predict_proba(testing.text)
] >= threshold_recall
training["classification_nb_recall"] = [
    proba[1] for proba in clf.predict_proba(training.text)
] >= threshold_recall

performance_statistics["Naive Bayes Threshold"] = classify.return_statistics(
    ground_truth=testing.relevant,
    scores=testing.score_nb,
    classification=testing.classification_nb_recall,
)

# %%
pipeline = Pipeline(
    [
        (
            "vect",
            CountVectorizer(
                strip_accents="unicode",
                stop_words=classify.stop_words,
                ngram_range=(1, 3),
                min_df=10,
            ),
        ),
        ("clf", LogisticRegression(penalty="l2")),
    ],
)
clf = pipeline.fit(training.text, training.relevant)

testing["classification_ridge"] = clf.predict(testing.text)
testing["score_ridge"] = [proba[1] for proba in clf.predict_proba(testing.text)]

training["classification_ridge"] = clf.predict(training.text)
training["score_ridge"] = [proba[1] for proba in clf.predict_proba(training.text)]

performance_statistics["Ridge Regression"] = classify.return_statistics(
    ground_truth=testing.relevant,
    scores=testing.score_ridge,
    classification=testing.classification_ridge,
)

cf_matrix = confusion_matrix(testing.relevant, testing.classification_ridge)
classify.create_plot_confusion_matrix(cf_matrix=cf_matrix)

pickle.dump(clf, open(start.TEMP_DIR + "model_ridge.sav", "wb"))


probas = [proba[1] for proba in clf.predict_proba(training.text)]
precisions, recalls, thresholds = precision_recall_curve(training.relevant, probas)
threshold_recall = thresholds[np.argmax(recalls >= 0.70)]


testing["classification_ridge_recall"] = [
    proba[1] for proba in clf.predict_proba(testing.text)
] >= threshold_recall
training["classification_ridge_recall"] = [
    proba[1] for proba in clf.predict_proba(training.text)
] >= threshold_recall


cf_matrix = confusion_matrix(testing.relevant, testing.classification_ridge_recall)
classify.create_plot_confusion_matrix(cf_matrix=cf_matrix)

performance_statistics["Ridge Regression Threshold"] = classify.return_statistics(
    ground_truth=testing.relevant,
    scores=testing.score_ridge,
    classification=testing.classification_ridge_recall,
)


# %%
training_spacy = pd.read_csv(start.TEMP_DIR + "training_spacy.csv")
testing_spacy = pd.read_csv(start.TEMP_DIR + "testing_spacy.csv")


training = training.merge(
    training_spacy[["unique_id", "classification"]].rename(
        columns={"classification": "score_spacy"}
    ),
    left_on="unique_id",
    right_on="unique_id",
)
training["classification_spacy"] = np.where(training.score_spacy > 0.5, 1, 0)

testing = testing.merge(
    testing_spacy[["unique_id", "classification"]].rename(
        columns={"classification": "score_spacy"}
    ),
    left_on="unique_id",
    right_on="unique_id",
)
testing["classification_spacy"] = np.where(testing.score_spacy > 0.5, 1, 0)

performance_statistics["SpaCy CNN"] = classify.return_statistics(
    ground_truth=testing.relevant,
    scores=testing.score_spacy,
    classification=testing.classification_spacy,
)

# %%
training["classification_rule"] = np.where(
    (training.score_spacy > 0.5)
    | (training.score_ridge > 0.5)
    | (training.score_svm > 0.5)
    | (training.score_sgd > 0.5)
    | (training.score_nb > 0.5)
    | (training.score_rf > 0.5),
    1,
    0,
)

testing["classification_rule"] = np.where(
    (testing.score_spacy > 0.5)
    | (testing.score_ridge > 0.5)
    | (testing.score_svm > 0.5)
    | (testing.score_sgd > 0.5)
    | (testing.score_nb > 0.5)
    | (testing.score_rf > 0.5),
    1,
    0,
)

testing["classification_count"] = (
    testing.classification_spacy
    + testing.classification_ridge
    + testing.classification_svm
    + testing.classification_sgd
    + testing.classification_nb
    + testing.classification_rf
)

performance_statistics["Ensemble"] = classify.return_statistics(
    ground_truth=testing.relevant,
    scores=testing.classification_count,
    classification=testing.classification_rule,
)
# %%
training.to_csv(start.TEMP_DIR + "training_models.csv")
testing.to_csv(start.TEMP_DIR + "testing_models.csv")

# %%
file_path = start.CLEAN_DIR + "performance_relevance.xlsx"
wb = load_workbook(file_path)
ws = wb.active

row = 2
for model in performance_statistics.keys():
    col = 1
    ws.cell(row=row, column=col).value = model
    col = col + 1
    ws.cell(row=row, column=col).value = performance_statistics[model]["accuracy"]
    col = col + 1
    ws.cell(row=row, column=col).value = performance_statistics[model]["precision"]
    col = col + 1
    ws.cell(row=row, column=col).value = performance_statistics[model]["recall"]
    col = col + 1
    ws.cell(row=row, column=col).value = performance_statistics[model]["auc"]
    col = col + 1
    ws.cell(row=row, column=col).value = performance_statistics[model]["specificity"]
    col = col + 1

    row = row + 1

wb.save(file_path)

# %%
# Proportion of relevant tweets in narrowed corpus
narrowed_test = testing[testing.classification_rule == 1]
print(len(narrowed_test[narrowed_test.relevant == 1]) / len(narrowed_test))


irrelevant_test = testing[testing.relevant == 0]
print(
    len(irrelevant_test[irrelevant_test.classification_rule == 0])
    / len(irrelevant_test)
)
