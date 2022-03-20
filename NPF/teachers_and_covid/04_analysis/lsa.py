# %%
import pandas as pd
import numpy as np

from NPF.teachers_and_covid import start
from NPF.library import process_text
from openpyxl import load_workbook

import matplotlib.pyplot as plt

df = pd.read_csv(start.CLEAN_DIR + "tweets_full.csv")
df = df.sample(50000)
df = df.set_index("unique_id")


# %%
# df["text_clean"] = [
#     process_text.process_text_nltk(
#         text,
#         lower_case=True,
#         remove_punct=True,
#         remove_stopwords=True,
#         lemma=True,
#         string_or_list="string",
#     )
#     for text in df.text
# ]

matrix = process_text.vectorize_text(
    df, "text", remove_stopwords=True, min_df=50, lemma=True, n_gram_range=(1, 1)
)

phrases = process_text.vectorize_text(
    df=df,
    text_col="text",
    remove_stopwords=False,
    min_df=50,
    n_gram_range=(2, 4),
)
columns = []
for phrase in phrases.columns:
    stop_phrase = False
    for stop in process_text.STOPLIST:
        if phrase.startswith(stop):
            stop_phrase = True
        if phrase.endswith(stop):
            stop_phrase = True
    if (not stop_phrase) and (phrase not in columns):
        columns.append(phrase)
phrases = phrases[columns]
phrases = phrases * 3

term_matrix = matrix.merge(phrases, left_index=True, right_index=True)

lsa = process_text.create_lsa_dfs(term_matrix, n_components=100, random_state=100)
lsa.matrix
lsa.word_weights

# %%
file_path = start.TEMP_DIR + "explore_lsa.xlsx"
wb = load_workbook(file_path)
ws = wb.active

col = 1
for component in list(range(0, 50)):
    words = lsa.word_weights[lsa.word_weights[component] > 0.1].sort_values(
        by=component, ascending=False
    )[component]
    ws.cell(row=1, column=col).value = component
    row = 2
    for word in words.index:
        ws.cell(row=row, column=col).value = word
        row = row + 1
    col = col + 1
    wb.save(file_path)


# %%
file_path = start.TEMP_DIR + "explore_lsa_subset.xlsx"
wb = load_workbook(file_path)
ws = wb.active

relevant_topics = [11, 13, 26, 38, 41, 44]

col = 1
for component in relevant_topics:
    words = lsa.word_weights[lsa.word_weights[component] > 0.1].sort_values(
        by=component, ascending=False
    )[component]
    ws.cell(row=1, column=col).value = component
    row = 2
    for word in words.index:
        ws.cell(row=row, column=col).value = word
        row = row + 1
    col = col + 1
    wb.save(file_path)

# %%
file_path = start.TEMP_DIR + "explore_lsa_tweets.xlsx"
wb = load_workbook(file_path)
ws = wb.active


col = 1
for component in relevant_topics:
    ws.cell(row=1, column=col).value = component

    tweet_indices = lsa.matrix.sort_values(by=component, ascending=False).head(5).index
    tweets_list = list(df.loc[tweet_indices]["text"])
    row = 2
    for tweet in tweets_list:
        ws.cell(row=row, column=col).value = tweet
        row = row + 1
    col = col + 1
    wb.save(file_path)

# %%
