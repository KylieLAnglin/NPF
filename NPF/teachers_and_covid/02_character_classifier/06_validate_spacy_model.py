# %%
# %%
from cgi import test
import re
import pandas as pd
import numpy as np
import pickle
from openpyxl import load_workbook


from NPF.teachers_and_covid import start
from NPF.library import classify


# %%
df = pd.read_csv(start.TEMP_DIR + "testing_characters_spacy.csv")
df = df[df.category.isin([1, 2, 3, 4])]

df["gold_hero"] = np.where(df.category == 1, 1, 0)
df["gold_villain"] = np.where(df.category == 2, 1, 0)
df["gold_victim"] = np.where(df.category == 3, 1, 0)

df["classification_hero"] = np.where(df.spacy_hero > 0.5, 1, 0)
df["classification_villain"] = np.where(df.spacy_villain > 0.5, 1, 0)
df["classification_victim"] = np.where(df.spacy_victim > 0.5, 1, 0)


# %%
performance_statistics = {}

performance_statistics["hero"] = classify.return_statistics(
    ground_truth=df.gold_hero,
    scores=df.spacy_hero,
    classification=df.classification_hero,
)

performance_statistics["villain"] = classify.return_statistics(
    ground_truth=df.gold_villain,
    scores=df.spacy_villain,
    classification=df.classification_villain,
)

performance_statistics["victim"] = classify.return_statistics(
    ground_truth=df.gold_victim,
    scores=df.spacy_victim,
    classification=df.classification_victim,
)

# %%

file_path = start.CLEAN_DIR + "performance_characters.xlsx"
wb = load_workbook(file_path)
ws = wb.active

row = 2
for model in performance_statistics.keys():
    col = 1
    ws.cell(row=row, column=col).value = model
    col = col + 1
    ws.cell(row=row, column=col).value = performance_statistics[model]["accuracy"]
    col = col + 1
    ws.cell(row=row, column=col).value = performance_statistics[model]["precision"]
    col = col + 1
    ws.cell(row=row, column=col).value = performance_statistics[model]["recall"]
    col = col + 1
    ws.cell(row=row, column=col).value = performance_statistics[model]["auc"]
    col = col + 1
    ws.cell(row=row, column=col).value = performance_statistics[model]["specificity"]
    col = col + 1

    row = row + 1

wb.save(file_path)

# %%
